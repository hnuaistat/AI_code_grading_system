import os
import re
import json
import uuid
import asyncio
import io
import unicodedata
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Dict, Optional, Set

import time
from fastapi import (
    FastAPI, Depends, HTTPException, status, UploadFile, File,
    Form, BackgroundTasks, Body, Request
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from sqlalchemy.orm import Session

load_dotenv()

import models
from database import engine, SessionLocal, get_db
from auth import (
    authenticate_user, create_access_token, get_current_user,
    get_password_hash, verify_password, ACCESS_TOKEN_EXPIRE_MINUTES, require_admin
)
from schemas import (
    Token, LoginRequest, RegisterRequest, GradingCriteria, GradingSession,
    StudentResult, SubjectCreate, SubjectResponse, HistorySessionItem, SubjectItemCreate,
    ProblemRevisionRequest, RevisionLogItem, SubjectUpdate, SubjectItemUpdate,
    DecomposeRequest, SessionSubjectItemUpdate, RegradeRequest,
    UpdateEmailRequest, ChangePasswordRequest
)
from services.notebook_service import (
    extract_notebooks_from_zip, parse_student_id_from_filename,
    parse_notebook, split_notebook_by_problems
)
from services.grading_service import grade_student_notebook, grade_student_problems
from services.llm_service import APIQuotaError, generate_rubric_with_ai, decompose_rubric_item_with_ai

app = FastAPI(title="Jupyter Notebook 자동 채점 시스템", version="2.0.0")

# CORS: FRONTEND_URL이 설정되면 해당 도메인 + 로컬 개발 서버만 허용
# 미설정 시 기존 동작(전체 허용) 유지 — 배포 환경변수에 FRONTEND_URL 설정 권장
_frontend_url = os.getenv("FRONTEND_URL", "").strip().rstrip("/")
_allowed_origins = (
    sorted({_frontend_url, "http://localhost:3000"}) if _frontend_url else ["*"]
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

# In-memory active sessions
grading_sessions: Dict[str, GradingSession] = {}

# 강제 중단 요청된 세션 id (background task 다음 iteration에서 종료)
cancelled_sessions: Set[str] = set()


# ─── Startup ───────────────────────────────────────────────────────────────────

def seed_database():
    """Create default users and subjects if they don't exist."""
    db = SessionLocal()
    try:
        # 보안: 시드 비밀번호는 코드에 하드코딩하지 않고 환경변수에서 읽는다.
        # 환경변수가 없으면 해당 계정은 생성하지 않음 (기존 DB의 계정은 영향 없음)
        seed_users = [
            {
                "username": "admin",
                "email": "admin@univ.ac.kr",
                "password": os.getenv("SEED_ADMIN_PASSWORD", ""),
                "role": "admin",
                "subjects": [],
            },
            {
                "username": "professor",
                "email": "professor@univ.ac.kr",
                "password": os.getenv("SEED_PROFESSOR_PASSWORD", ""),
                "role": "professor",
                "subjects": [("알고리즘", "CS101"), ("자료구조", "CS102")],
            },
            {
                "username": "prof_kim",
                "email": "kim.prof@univ.ac.kr",
                "password": os.getenv("SEED_PROF_KIM_PASSWORD", ""),
                "role": "professor",
                "subjects": [("데이터베이스", "DB201"), ("운영체제", "OS202")],
            },
        ]
        for u in seed_users:
            if not u["password"]:
                continue
            existing = db.query(models.User).filter(models.User.username == u["username"]).first()
            if not existing:
                user = models.User(
                    username=u["username"],
                    email=u["email"],
                    hashed_password=get_password_hash(u["password"]),
                    role=u["role"],
                )
                db.add(user)
                db.flush()
                for name, code in u["subjects"]:
                    db.add(models.Subject(name=name, code=code, user_id=user.id))
        db.commit()

        # 기본 시스템 설정 시드
        default_settings = {
            "openai_api_key": os.getenv("OPENAI_API_KEY", ""),
            "fireworks_api_key": os.getenv("FIREWORKS_API_KEY", ""),
            "llm_model": "gpt-4o-mini",
            "base_system_prompt": "",
            "max_upload_size_mb": "50",
        }
        for key, value in default_settings.items():
            existing = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
            if not existing:
                db.add(models.SystemSetting(key=key, value=value))
        db.commit()
    finally:
        db.close()


def _migrate_add_columns():
    """기존 테이블에 누락된 컬럼을 ALTER TABLE로 추가 (PostgreSQL/SQLite 호환)."""
    from sqlalchemy import text, inspect
    inspector = inspect(engine)
    migrations = [
        ("grading_sessions_db", "grading_model", "VARCHAR(200)"),
        ("grading_sessions_db", "criteria_json", "TEXT"),
        ("grading_sessions_db", "answer_problems_json", "TEXT"),
        ("grading_sessions_db", "regraded_from", "VARCHAR(36)"),
    ]
    with engine.begin() as conn:
        for table_name, col_name, col_type in migrations:
            try:
                if not inspector.has_table(table_name):
                    continue
                existing_cols = {c["name"] for c in inspector.get_columns(table_name)}
                if col_name not in existing_cols:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_type}"))
                    print(f"[Migration] Added column {table_name}.{col_name}")
            except Exception as e:
                print(f"[Migration] {table_name}.{col_name} 추가 실패 (이미 존재할 수 있음): {e}")


@app.on_event("startup")
async def startup():
    try:
        models.Base.metadata.create_all(bind=engine)
        _migrate_add_columns()
        seed_database()
    except Exception as e:
        print(f"Warning: Database initialization failed: {str(e)}")
        print("App will continue running, but with empty database.")


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/debug/openai-test")
async def debug_openai_test(admin=Depends(require_admin)):
    import httpx
    api_key = os.getenv("OPENAI_API_KEY", "")
    base_url = os.getenv("OPENAI_BASE_URL") or "https://api.openai.com/v1"
    results = {
        "api_key_set": bool(api_key),
        "api_key_prefix": api_key[:8] if api_key else "",
        "base_url": base_url,
    }
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            r = await client.get(
                f"{base_url}/models",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            results["httpx_status"] = r.status_code
            results["httpx_body"] = r.text[:300]
    except Exception as e:
        results["httpx_error"] = f"{type(e).__name__}: {str(e)}"
    return results


# ─── Auth ──────────────────────────────────────────────────────────────────────

# 로그인 무차별 대입 방지: (IP, 아이디)별 실패 5회 → 5분 잠금 (인메모리)
_login_fails: Dict[str, Dict[str, float]] = {}
LOGIN_MAX_FAILS = 5
LOGIN_LOCK_SECONDS = 300


@app.post("/auth/login", response_model=Token)
async def login(request: LoginRequest, req: Request, db: Session = Depends(get_db)):
    client_ip = req.client.host if req.client else "unknown"
    fail_key = f"{client_ip}|{request.username}"
    now = time.time()

    rec = _login_fails.get(fail_key)
    if rec and rec["count"] >= LOGIN_MAX_FAILS:
        remaining = LOGIN_LOCK_SECONDS - (now - rec["last"])
        if remaining > 0:
            raise HTTPException(
                status_code=429,
                detail=f"로그인 시도가 너무 많습니다. {int(remaining // 60) + 1}분 후 다시 시도해주세요",
            )
        del _login_fails[fail_key]  # 잠금 시간 경과 → 초기화

    user = authenticate_user(db, request.username, request.password)
    if not user:
        rec = _login_fails.setdefault(fail_key, {"count": 0, "last": now})
        rec["count"] += 1
        rec["last"] = now
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="아이디 또는 비밀번호가 올바르지 않습니다",
            headers={"WWW-Authenticate": "Bearer"},
        )
    _login_fails.pop(fail_key, None)  # 성공 시 실패 기록 초기화
    token = create_access_token(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return Token(access_token=token, token_type="bearer")


@app.post("/auth/register")
async def register(request: RegisterRequest, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.username == request.username).first():
        raise HTTPException(status_code=400, detail="이미 사용 중인 아이디입니다")
    if db.query(models.User).filter(models.User.email == request.email).first():
        raise HTTPException(status_code=400, detail="이미 사용 중인 이메일입니다")
    if len(request.password) < 6:
        raise HTTPException(status_code=400, detail="비밀번호는 6자 이상이어야 합니다")

    user = models.User(
        username=request.username,
        email=request.email,
        hashed_password=get_password_hash(request.password),
        role="professor",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"message": "회원가입이 완료되었습니다", "username": user.username}


@app.get("/auth/me")
async def get_me(current_user=Depends(get_current_user)):
    return current_user


@app.patch("/auth/me")
async def update_me(
    request: UpdateEmailRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """이메일 변경 (아이디/역할은 변경 불가)."""
    email = request.email.strip()
    if not email or "@" not in email or "." not in email.split("@")[-1]:
        raise HTTPException(status_code=400, detail="올바른 이메일 형식이 아닙니다")
    dup = db.query(models.User).filter(
        models.User.email == email,
        models.User.id != current_user["id"]
    ).first()
    if dup:
        raise HTTPException(status_code=400, detail="이미 사용 중인 이메일입니다")

    user = db.query(models.User).filter(models.User.id == current_user["id"]).first()
    user.email = email
    db.commit()
    return {"id": user.id, "username": user.username, "email": user.email, "role": user.role}


@app.post("/auth/change-password")
async def change_password(
    request: ChangePasswordRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """비밀번호 변경 — 현재 비밀번호 확인 필수."""
    user = db.query(models.User).filter(models.User.id == current_user["id"]).first()
    if not verify_password(request.current_password, user.hashed_password):
        raise HTTPException(status_code=400, detail="현재 비밀번호가 올바르지 않습니다")
    if len(request.new_password) < 6:
        raise HTTPException(status_code=400, detail="새 비밀번호는 6자 이상이어야 합니다")
    if request.new_password == request.current_password:
        raise HTTPException(status_code=400, detail="새 비밀번호가 현재 비밀번호와 같습니다")

    user.hashed_password = get_password_hash(request.new_password)
    db.commit()
    return {"message": "비밀번호가 변경되었습니다"}


# ─── Subjects ──────────────────────────────────────────────────────────────────

@app.get("/subjects")
async def list_subjects(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    subjects = (
        db.query(models.Subject)
        .filter(models.Subject.user_id == current_user["id"])
        .order_by(models.Subject.created_at)
        .all()
    )
    result = []
    for s in subjects:
        count = db.query(models.GradingSessionDB).filter(
            models.GradingSessionDB.subject_id == s.id
        ).count()
        items = [{"id": item.id, "name": item.name, "created_at": item.created_at.isoformat()} for item in s.items]
        result.append({
            "id": s.id,
            "name": s.name,
            "code": s.code,
            "session_count": count,
            "items": items,
            "created_at": s.created_at.isoformat(),
        })
    return result


@app.post("/subjects")
async def create_subject(
    body: SubjectCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    subject = models.Subject(
        name=body.name,
        code=body.code,
        user_id=current_user["id"],
    )
    db.add(subject)
    db.commit()
    db.refresh(subject)
    return {"id": subject.id, "name": subject.name, "code": subject.code, "session_count": 0,
            "items": [], "created_at": subject.created_at.isoformat()}


@app.get("/subjects/{subject_id}")
async def get_subject(
    subject_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.user_id == current_user["id"]
    ).first()
    if not subject:
        raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다")

    items = [{"id": item.id, "name": item.name, "created_at": item.created_at.isoformat()} for item in subject.items]
    count = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.subject_id == subject.id
    ).count()
    return {
        "id": subject.id,
        "name": subject.name,
        "code": subject.code,
        "session_count": count,
        "items": items,
        "created_at": subject.created_at.isoformat(),
    }


@app.post("/subjects/{subject_id}/items")
async def create_subject_item(
    subject_id: int,
    body: SubjectItemCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.user_id == current_user["id"]
    ).first()
    if not subject:
        raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다")

    item = models.SubjectItem(subject_id=subject_id, name=body.name)
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id, "name": item.name, "created_at": item.created_at.isoformat()}


@app.put("/subjects/{subject_id}")
async def update_subject(
    subject_id: int,
    body: SubjectUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.user_id == current_user["id"]
    ).first()
    if not subject:
        raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다")

    if body.name is not None:
        subject.name = body.name
    if body.code is not None:
        subject.code = body.code if body.code.strip() else None
    db.commit()
    db.refresh(subject)
    count = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.subject_id == subject.id
    ).count()
    items = [{"id": item.id, "name": item.name, "created_at": item.created_at.isoformat()} for item in subject.items]
    return {"id": subject.id, "name": subject.name, "code": subject.code,
            "session_count": count, "items": items, "created_at": subject.created_at.isoformat()}


@app.put("/subjects/{subject_id}/items/{item_id}")
async def update_subject_item(
    subject_id: int,
    item_id: int,
    body: SubjectItemUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.user_id == current_user["id"]
    ).first()
    if not subject:
        raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다")

    item = db.query(models.SubjectItem).filter(
        models.SubjectItem.id == item_id,
        models.SubjectItem.subject_id == subject_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다")

    item.name = body.name
    db.commit()
    db.refresh(item)
    return {"id": item.id, "name": item.name, "created_at": item.created_at.isoformat()}


@app.delete("/subjects/{subject_id}/items/{item_id}")
async def delete_subject_item(
    subject_id: int,
    item_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.user_id == current_user["id"]
    ).first()
    if not subject:
        raise HTTPException(status_code=404, detail="과목을 찾을 수 없습니다")

    item = db.query(models.SubjectItem).filter(
        models.SubjectItem.id == item_id,
        models.SubjectItem.subject_id == subject_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="항목을 찾을 수 없습니다")

    db.delete(item)
    db.commit()
    return {"message": "항목이 삭제되었습니다"}


# ─── LLM 모델 목록 ─────────────────────────────────────────────────────────────

@app.get("/grading/available-models")
async def list_available_models(current_user=Depends(get_current_user)):
    """채점에 사용 가능한 LLM 모델 목록 반환. 환경변수에 키가 있는 provider만 노출."""
    from services.llm_service import AVAILABLE_MODELS, DEFAULT_MODEL
    has_openai = bool(os.getenv("OPENAI_API_KEY", "").strip())
    has_fireworks = bool(os.getenv("FIREWORKS_API_KEY", "").strip())
    available = []
    for m in AVAILABLE_MODELS:
        provider = m.get("provider", "openai")
        if provider == "openai" and not has_openai:
            continue
        if provider == "fireworks" and not has_fireworks:
            continue
        available.append(m)
    return {"models": available, "default": DEFAULT_MODEL}


# ─── Rubric Generation ────────────────────────────────────────────────────────

@app.post("/grading/generate-rubric")
async def generate_rubric(
    answer_notebook: UploadFile = File(...),
    total_score: float = Form(100.0),
    exam_title: str = Form(""),
    current_user=Depends(get_current_user),
):
    """정답 노트북을 분석하여 루브릭 JSON을 자동 생성합니다."""
    answer_bytes = await answer_notebook.read()

    try:
        nb = parse_notebook(answer_bytes)
        problems = split_notebook_by_problems(nb)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"노트북 파싱 오류: {str(e)}")

    if not problems:
        raise HTTPException(status_code=400, detail="노트북에서 문제를 찾을 수 없습니다. 마크다운 셀에 Q1, Q2 등의 문제 마커가 필요합니다.")

    try:
        rubric = await generate_rubric_with_ai(
            answer_problems=problems,
            total_score=total_score,
            exam_title=exam_title,
        )
        return rubric
    except APIQuotaError:
        raise HTTPException(status_code=429, detail="OpenAI API 사용량이 초과되었습니다.")
    except ValueError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"루브릭 생성 실패: {str(e)}")


# ─── Grading ───────────────────────────────────────────────────────────────────

@app.post("/grading/start")
async def start_grading(
    background_tasks: BackgroundTasks,
    answer_notebook: UploadFile = File(...),
    student_zip: UploadFile = File(...),
    criteria_file: UploadFile = File(...),
    subject_id: Optional[int] = Form(None),
    subject_item_id: Optional[int] = Form(None),
    grading_model: Optional[str] = Form(None),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    answer_bytes = await answer_notebook.read()
    student_bytes = await student_zip.read()
    criteria_bytes = await criteria_file.read()

    try:
        criteria_data = json.loads(criteria_bytes.decode('utf-8'))
        criteria = GradingCriteria(**criteria_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"채점 기준 파일 파싱 오류: {str(e)}")

    try:
        filename = student_zip.filename or ""
        if filename.lower().endswith('.ipynb'):
            student_notebooks = [(filename, student_bytes)]
        else:
            student_notebooks = extract_notebooks_from_zip(student_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"학생 제출물 처리 오류: {str(e)}")

    if not student_notebooks:
        raise HTTPException(status_code=400, detail="제출물에서 .ipynb 파일을 찾을 수 없습니다")

    session_id = str(uuid.uuid4())
    session = GradingSession(
        session_id=session_id,
        status="pending",
        progress=0.0,
        total_students=len(student_notebooks),
        processed_students=0,
        results=[]
    )
    grading_sessions[session_id] = session

    # 모델 ID 정규화 (None/빈문자 → 기본값)
    from services.llm_service import DEFAULT_MODEL, AVAILABLE_MODELS
    valid_model_ids = {m["id"] for m in AVAILABLE_MODELS}
    if not grading_model or grading_model not in valid_model_ids:
        grading_model = DEFAULT_MODEL

    # 재채점용 입력 저장: 루브릭 원본 + 문항별 정답 데이터 (이미지 제외, 텍스트만)
    answer_problems_json = None
    try:
        answer_nb = parse_notebook(answer_bytes)
        answer_problems_data = split_notebook_by_problems(answer_nb)
        answer_problems_json = json.dumps(
            _strip_problem_images(answer_problems_data), ensure_ascii=False, default=str
        )
    except Exception as e:
        print(f"[WARNING] 정답 데이터 저장 실패 (재채점 불가 세션): {e}")

    # Persist initial record to DB
    db_record = models.GradingSessionDB(
        id=session_id,
        subject_id=subject_id,
        subject_item_id=subject_item_id,
        user_id=current_user["id"],
        status="running",
        total_students=len(student_notebooks),
        processed_students=0,
        grading_model=grading_model,
        criteria_json=json.dumps(criteria_data, ensure_ascii=False),
        answer_problems_json=answer_problems_json,
    )
    db.add(db_record)
    db.commit()

    background_tasks.add_task(
        run_grading_session,
        session_id, answer_bytes, student_notebooks, criteria,
        subject_id, current_user["id"], grading_model
    )

    return {"session_id": session_id, "total_students": len(student_notebooks)}


def _extract_student_info(nb_content: bytes) -> tuple[str, str]:
    """노트북 첫 셀에서 '# 학번 :'과 '# 이름 :' 추출. (학번, 이름) 튜플 반환."""
    try:
        from services.notebook_service import parse_notebook
        nb = parse_notebook(nb_content)
        if nb.cells:
            src = nb.cells[0].source if isinstance(nb.cells[0].source, str) else ''.join(nb.cells[0].source)
            student_id, name = "", ""
            for line in src.split('\n'):
                line_stripped = line.strip().lstrip('#').strip()
                if line_stripped.startswith('학번'):
                    student_id = line_stripped.split(':', 1)[-1].strip()
                elif line_stripped.startswith('이름'):
                    name = line_stripped.split(':', 1)[-1].strip()
            return student_id, name
    except Exception:
        pass
    return "", ""


def _dump_strip_images(r: dict) -> dict:
    """DB 저장용: code_cells/preamble_cells에서 base64 이미지만 제거, 텍스트는 보존"""
    for p in r.get("problems", []):
        for cell in p.get("code_cells", []):
            for out in cell.get("outputs", []):
                out["image"] = None
        for cell in p.get("preamble_cells", []):
            for out in cell.get("outputs", []):
                out["image"] = None
    return r


def _strip_problem_images(problems: dict) -> dict:
    """재채점용 정답 데이터 저장 시 셀 출력에서 이미지 제거 (텍스트만 보존)"""
    stripped = {}
    for pid, info in problems.items():
        cells = []
        for c in info.get("cells", []):
            cell = dict(c)
            outputs = []
            for o in cell.get("outputs", []) or []:
                out = dict(o)
                if isinstance(out.get("data"), dict):
                    out["data"] = {"text/plain": out["data"].get("text/plain", "")}
                outputs.append(out)
            cell["outputs"] = outputs
            cells.append(cell)
        stripped[pid] = {"description": info.get("description", ""), "cells": cells}
    return stripped


def _pid_num(pid) -> int:
    """'Q1', '문제1' 등에서 문제 번호 추출"""
    if isinstance(pid, int):
        return pid
    digits = re.sub(r"\D", "", str(pid))
    return int(digits) if digits else 0


def _nbcell_to_raw(c: dict) -> dict:
    """results_json에 저장된 NotebookCell을 채점 파이프라인의 원본 셀 형태로 복원"""
    if c.get("cell_type") == "markdown":
        return {
            "cell_type": "markdown",
            "source": c.get("source", ""),
            "outputs": [],
            "is_student_answer": c.get("is_student_answer", False),
        }
    outputs = []
    for o in c.get("outputs", []) or []:
        otype = o.get("output_type", "")
        text = o.get("text") or ""
        if otype == "stream":
            outputs.append({"output_type": "stream", "text": text})
        elif otype in ("execute_result", "display_data"):
            outputs.append({"output_type": otype, "data": {"text/plain": text}})
        elif otype == "error":
            # 저장 형태: "ename: evalue\n(traceback)" → 첫 줄에서 복원
            first_line = text.split("\n", 1)[0]
            ename, _, evalue = first_line.partition(": ")
            outputs.append({"output_type": "error", "ename": ename, "evalue": evalue, "traceback": []})
    return {"cell_type": "code", "source": c.get("source", ""), "outputs": outputs}


def _reconstruct_student_problems(student: dict) -> dict:
    """results_json의 학생 결과에서 문항별 셀 데이터를 복원 (재채점 입력용)"""
    problems = {}
    for p in student.get("problems", []):
        cells = [_nbcell_to_raw(c) for c in (p.get("code_cells") or [])]
        problems[_pid_num(p.get("problem_id"))] = {
            "description": p.get("problem_description", ""),
            "cells": cells,
        }
        preamble = p.get("preamble_cells") or []
        if preamble and 0 not in problems:
            problems[0] = {"description": "", "cells": [_nbcell_to_raw(c) for c in preamble]}
    return problems


def _persist_session_results(session_id: str, session: GradingSession, session_total_tokens: int):
    """채점/재채점 완료 후 세션 결과를 DB에 저장 (공용)"""
    db = SessionLocal()
    try:
        results_data = [_dump_strip_images(r.model_dump()) for r in session.results]
        db_record = db.query(models.GradingSessionDB).filter(
            models.GradingSessionDB.id == session_id
        ).first()
        if db_record:
            # cancel endpoint가 먼저 DB를 cancelled로 마킹했을 수 있으므로 유지
            final_status = "cancelled" if session_id in cancelled_sessions else session.status
            db_record.status = final_status
            db_record.progress = session.progress
            db_record.processed_students = session.processed_students
            db_record.results_json = json.dumps(results_data, ensure_ascii=False)
            db_record.error = session.error
            db_record.tokens_used = (db_record.tokens_used or 0) + session_total_tokens
            if final_status == "completed":
                db_record.completed_at = datetime.utcnow()
            db.commit()
    finally:
        db.close()
        cancelled_sessions.discard(session_id)


async def run_grading_session(
    session_id: str,
    answer_bytes: bytes,
    student_notebooks: list,
    criteria: GradingCriteria,
    subject_id: Optional[int] = None,
    user_id: Optional[int] = None,
    grading_model: Optional[str] = None,
):
    session = grading_sessions[session_id]
    session.status = "running"
    total = len(student_notebooks)

    quota_exceeded = False
    cancelled = False
    session_total_tokens = 0
    for i, (filename, content) in enumerate(student_notebooks):
        if session_id in cancelled_sessions:
            cancelled = True
            session.processed_students = i
            session.progress = (i / total) * 100 if total else 0
            break
        session.current_student = filename
        try:
            problem_results, error, student_tokens = await grade_student_notebook(
                student_nb_content=content,
                answer_nb_content=answer_bytes,
                criteria=criteria,
                execute=False,
                model=grading_model,
            )
            total_score = sum(p.obtained_score for p in problem_results)
            max_total = sum(p.full_score for p in problem_results)

            # 노트북에서 학번/이름 추출
            nb_student_id, nb_student_name = _extract_student_info(content)

            student_result = StudentResult(
                filename=filename,
                student_id=nb_student_id or parse_student_id_from_filename(filename),
                student_name=nb_student_name,
                total_score=total_score,
                max_total_score=max_total,
                problems=problem_results,
                error=error
            )
            session.results.append(student_result)
            session_total_tokens += student_tokens
        except APIQuotaError:
            quota_exceeded = True
            session.processed_students = i
            session.progress = (i / total) * 100
            break
        except Exception as e:
            nb_student_id, nb_student_name = _extract_student_info(content)
            session.results.append(StudentResult(
                filename=filename,
                student_id=nb_student_id or parse_student_id_from_filename(filename),
                student_name=nb_student_name,
                total_score=0,
                max_total_score=sum(p.full_score for p in criteria.problems),
                problems=[],
                error=str(e)
            ))

        session.processed_students = i + 1
        session.progress = ((i + 1) / total) * 100

    if cancelled:
        session.status = "cancelled"
        session.current_student = None
        session.error = "사용자 요청으로 채점이 중단되었습니다"
    elif quota_exceeded:
        session.status = "quota_exceeded"
        session.current_student = None
        session.error = "OpenAI API 사용량이 초과되었습니다. API를 충전한 후 이어서 채점하세요."
    else:
        session.status = "completed"
        session.progress = 100.0
        session.current_student = None

    # Persist to DB
    _persist_session_results(session_id, session, session_total_tokens)


async def run_regrade_session(
    session_id: str,
    answer_problems: dict,
    students: list,
    criteria: GradingCriteria,
    grading_model: str,
):
    """저장된 입력(정답/루브릭/학생 데이터)으로 다른 모델 재채점 실행"""
    session = grading_sessions[session_id]
    session.status = "running"
    total = len(students)

    quota_exceeded = False
    cancelled = False
    session_total_tokens = 0
    for i, stu in enumerate(students):
        if session_id in cancelled_sessions:
            cancelled = True
            session.processed_students = i
            session.progress = (i / total) * 100 if total else 0
            break
        filename = stu.get("filename", "")
        session.current_student = filename
        try:
            student_problems = _reconstruct_student_problems(stu)
            problem_results, error, student_tokens = await grade_student_problems(
                student_problems=student_problems,
                answer_problems=answer_problems,
                criteria=criteria,
                model=grading_model,
            )
            total_score = sum(p.obtained_score for p in problem_results)
            max_total = sum(p.full_score for p in problem_results)
            session.results.append(StudentResult(
                filename=filename,
                student_id=stu.get("student_id", ""),
                student_name=stu.get("student_name", ""),
                total_score=total_score,
                max_total_score=max_total,
                problems=problem_results,
                error=error,
            ))
            session_total_tokens += student_tokens
        except APIQuotaError:
            quota_exceeded = True
            session.processed_students = i
            session.progress = (i / total) * 100
            break
        except Exception as e:
            session.results.append(StudentResult(
                filename=filename,
                student_id=stu.get("student_id", ""),
                student_name=stu.get("student_name", ""),
                total_score=0,
                max_total_score=sum(p.full_score for p in criteria.problems),
                problems=[],
                error=str(e),
            ))

        session.processed_students = i + 1
        session.progress = ((i + 1) / total) * 100

    if cancelled:
        session.status = "cancelled"
        session.current_student = None
        session.error = "사용자 요청으로 채점이 중단되었습니다"
    elif quota_exceeded:
        session.status = "quota_exceeded"
        session.current_student = None
        session.error = "API 사용량이 초과되었습니다. 충전 후 다시 재채점하세요."
    else:
        session.status = "completed"
        session.progress = 100.0
        session.current_student = None

    _persist_session_results(session_id, session, session_total_tokens)


@app.post("/grading/resume/{session_id}")
async def resume_grading(
    session_id: str,
    background_tasks: BackgroundTasks,
    answer_notebook: UploadFile = File(...),
    student_zip: UploadFile = File(...),
    criteria_file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record or db_record.status != "quota_exceeded":
        raise HTTPException(status_code=400, detail="이어서 채점할 수 없는 세션입니다 (quota_exceeded 상태가 아님)")

    answer_bytes = await answer_notebook.read()
    student_bytes = await student_zip.read()
    criteria_bytes = await criteria_file.read()

    try:
        criteria_data = json.loads(criteria_bytes.decode('utf-8'))
        criteria = GradingCriteria(**criteria_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"채점 기준 파일 파싱 오류: {str(e)}")

    try:
        filename = student_zip.filename or ""
        if filename.lower().endswith('.ipynb'):
            all_notebooks = [(filename, student_bytes)]
        else:
            all_notebooks = extract_notebooks_from_zip(student_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"학생 제출물 처리 오류: {str(e)}")

    # 이미 채점된 학생 파일명 수집
    already_done = set()
    if db_record.results_json:
        try:
            already_done = {r['filename'] for r in json.loads(db_record.results_json)}
        except Exception:
            pass

    remaining = [(f, c) for f, c in all_notebooks if f not in already_done]
    if not remaining:
        raise HTTPException(status_code=400, detail="이어서 채점할 학생이 없습니다 (모두 완료됨)")

    # 기존 결과 복원 후 세션 갱신
    existing_results = []
    if db_record.results_json:
        try:
            existing_results = [StudentResult(**r) for r in json.loads(db_record.results_json)]
        except Exception:
            pass

    new_total = len(existing_results) + len(remaining)
    session = GradingSession(
        session_id=session_id,
        status="running",
        progress=db_record.progress,
        total_students=new_total,
        processed_students=len(existing_results),
        results=existing_results,
        error=None,
    )
    grading_sessions[session_id] = session

    db_record.status = "running"
    db_record.total_students = new_total
    db_record.error = None
    db.commit()

    background_tasks.add_task(
        run_grading_session,
        session_id, answer_bytes, remaining, criteria,
        db_record.subject_id, current_user["id"]
    )

    return {"session_id": session_id, "remaining_students": len(remaining)}


@app.get("/grading/session/{session_id}")
async def get_session(
    session_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    def _strip(r: dict) -> dict:
        """code_cells/preamble_cells 제거 (base64 이미지 등 대용량 데이터 제외)"""
        for p in r.get("problems", []):
            p["code_cells"] = []
            p["preamble_cells"] = []
        return r

    session = grading_sessions.get(session_id)
    if session:
        # 인메모리 세션: 수동 직렬화로 대용량 셀 제외
        results_stripped = []
        for r in session.results:
            results_stripped.append({
                "filename": r.filename,
                "student_id": r.student_id,
                "student_name": r.student_name,
                "total_score": r.total_score,
                "max_total_score": r.max_total_score,
                "error": r.error,
                "problems": [{
                    "problem_id": p.problem_id,
                    "full_score": p.full_score,
                    "obtained_score": p.obtained_score,
                    "output_match": p.output_match,
                    "partial_scores": [ps.dict() for ps in p.partial_scores],
                    "ai_feedback": p.ai_feedback,
                    "code_cells": [],
                    "preamble_cells": [],
                    "problem_description": p.problem_description,
                    "professor_feedback": p.professor_feedback,
                    "is_revised": p.is_revised,
                    "revised_at": p.revised_at,
                    "has_ai_error": p.has_ai_error,
                    "has_partial_score": p.has_partial_score,
                } for p in r.problems],
            })
        # 인메모리 세션: DB에서 subject/model 정보 보완
        db_rec = db.query(models.GradingSessionDB).filter(
            models.GradingSessionDB.id == session_id
        ).first()
        subject_name = db_rec.subject.name if db_rec and db_rec.subject else None
        subject_code = db_rec.subject.code if db_rec and db_rec.subject else None
        subject_item_name = None
        if db_rec and db_rec.subject_item_id:
            si = db.query(models.SubjectItem).filter(models.SubjectItem.id == db_rec.subject_item_id).first()
            subject_item_name = si.name if si else None
        grading_model = db_rec.grading_model if db_rec else None
        created_at = _to_kst(db_rec.created_at) if db_rec else None
        completed_at = _to_kst(db_rec.completed_at) if db_rec and db_rec.completed_at else None

        # cancel 요청이 들어왔으면 즉시 'cancelled' 로 표시 (메모리 루프가 다음 iteration에서 종료됨)
        is_cancelled = session_id in cancelled_sessions
        return {
            "session_id": session.session_id,
            "status": "cancelled" if is_cancelled else session.status,
            "progress": session.progress,
            "current_student": None if is_cancelled else session.current_student,
            "total_students": session.total_students,
            "processed_students": session.processed_students,
            "results": results_stripped,
            "error": "사용자 요청으로 채점이 중단되었습니다" if is_cancelled else session.error,
            "subject_name": subject_name,
            "subject_code": subject_code,
            "subject_item_name": subject_item_name,
            "grading_model": grading_model,
            "regraded_from": db_rec.regraded_from if db_rec else None,
            "created_at": created_at,
            "completed_at": completed_at,
        }

    # DB에서 로드
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    results = []
    if db_record.results_json:
        try:
            results = [_strip(r) for r in json.loads(db_record.results_json)]
        except Exception:
            pass

    subject_item_name = None
    if db_record.subject_item_id:
        si = db.query(models.SubjectItem).filter(models.SubjectItem.id == db_record.subject_item_id).first()
        subject_item_name = si.name if si else None

    return {
        "session_id": session_id,
        "status": db_record.status,
        "progress": 100.0 if db_record.status == "completed" else db_record.progress,
        "current_student": None,
        "total_students": db_record.total_students,
        "processed_students": db_record.processed_students,
        "results": results,
        "error": db_record.error,
        "subject_name": db_record.subject.name if db_record.subject else None,
        "subject_code": db_record.subject.code if db_record.subject else None,
        "subject_item_name": subject_item_name,
        "grading_model": db_record.grading_model,
        "regraded_from": db_record.regraded_from,
        "created_at": _to_kst(db_record.created_at),
        "completed_at": _to_kst(db_record.completed_at) if db_record.completed_at else None,
    }


@app.get("/grading/session/{session_id}/student")
async def get_student_detail(
    session_id: str,
    filename: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """특정 학생의 전체 데이터(code_cells/preamble_cells 포함) 조회 — 모달 열 때만 호출"""
    mem_session = grading_sessions.get(session_id)
    if mem_session:
        student = next((r for r in mem_session.results if r.filename == filename), None)
        if student:
            return student

    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record or not db_record.results_json:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    raw = json.loads(db_record.results_json)
    student_data = next((r for r in raw if r.get("filename") == filename), None)
    if not student_data:
        raise HTTPException(status_code=404, detail="학생 데이터를 찾을 수 없습니다")

    return student_data


def _to_kst(dt: datetime) -> str:
    """UTC datetime을 KST(Asia/Seoul)로 변환하여 ISO 문자열 반환."""
    if not dt:
        return ""
    kst = timezone(timedelta(hours=9))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(kst).isoformat()


@app.get("/grading/history")
async def get_history(
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    from services.llm_service import AVAILABLE_MODELS
    # model ID → label 매핑
    model_label_map = {m["id"]: m["label"] for m in AVAILABLE_MODELS}

    records = (
        db.query(models.GradingSessionDB)
        .filter(models.GradingSessionDB.user_id == current_user["id"])
        .order_by(models.GradingSessionDB.created_at.desc())
        .all()
    )
    result = []
    for r in records:
        subject_item_name = None
        if r.subject_item_id:
            subject_item = db.query(models.SubjectItem).filter(
                models.SubjectItem.id == r.subject_item_id
            ).first()
            subject_item_name = subject_item.name if subject_item else None

        result.append({
            "session_id": r.id,
            "subject_id": r.subject_id,
            "subject_name": r.subject.name if r.subject else None,
            "subject_code": r.subject.code if r.subject else None,
            "subject_item_id": r.subject_item_id,
            "subject_item_name": subject_item_name,
            "status": r.status,
            "total_students": r.total_students,
            "processed_students": r.processed_students,
            "grading_model": r.grading_model,
            "grading_model_label": model_label_map.get(r.grading_model, r.grading_model),
            "regraded_from": r.regraded_from,
            "can_regrade": bool(r.criteria_json and r.answer_problems_json and r.results_json),
            "created_at": _to_kst(r.created_at),
            "completed_at": _to_kst(r.completed_at) if r.completed_at else None,
        })
    return result


@app.patch("/grading/session/{session_id}/subject-item")
async def update_session_subject_item(
    session_id: str,
    body: SessionSubjectItemUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """채점 완료된 세션의 세부 항목을 추가/수정한다.
    이름이 같은 항목이 과목에 이미 있으면 재사용, 없으면 새로 생성. 빈 문자열이면 해제."""
    record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id,
        models.GradingSessionDB.user_id == current_user["id"]
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    name = body.subject_item_name.strip()
    if not name:
        record.subject_item_id = None
        db.commit()
        return {"subject_item_id": None, "subject_item_name": None}

    if not record.subject_id:
        raise HTTPException(status_code=400, detail="과목이 지정되지 않은 세션은 세부 항목을 설정할 수 없습니다")

    item = db.query(models.SubjectItem).filter(
        models.SubjectItem.subject_id == record.subject_id,
        models.SubjectItem.name == name
    ).first()
    if not item:
        item = models.SubjectItem(subject_id=record.subject_id, name=name)
        db.add(item)
        db.flush()

    record.subject_item_id = item.id
    db.commit()
    return {"subject_item_id": item.id, "subject_item_name": item.name}


@app.post("/grading/session/{session_id}/regrade")
async def regrade_session(
    session_id: str,
    body: RegradeRequest,
    background_tasks: BackgroundTasks,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """저장된 입력(루브릭+정답+학생 데이터)으로 다른 모델 재채점.
    기존 세션은 보존하고 새 세션을 생성하며, regraded_from으로 원본(루트)을 참조."""
    orig = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id,
        models.GradingSessionDB.user_id == current_user["id"]
    ).first()
    if not orig:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")
    if orig.status != "completed":
        raise HTTPException(status_code=400, detail="완료된 세션만 재채점할 수 있습니다")
    if not (orig.criteria_json and orig.answer_problems_json and orig.results_json):
        raise HTTPException(status_code=400, detail="재채점 데이터가 저장되지 않은 세션입니다 (기능 추가 이전에 채점됨)")

    from services.llm_service import AVAILABLE_MODELS
    valid_model_ids = {m["id"] for m in AVAILABLE_MODELS}
    if body.grading_model not in valid_model_ids:
        raise HTTPException(status_code=400, detail="지원하지 않는 모델입니다")
    if body.grading_model == orig.grading_model:
        raise HTTPException(status_code=400, detail="이미 이 모델로 채점된 세션입니다. 다른 모델을 선택하세요.")

    try:
        criteria = GradingCriteria(**json.loads(orig.criteria_json))
        answer_problems = {int(k): v for k, v in json.loads(orig.answer_problems_json).items()}
        students = json.loads(orig.results_json)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"저장된 채점 데이터 복원 실패: {str(e)}")
    if not students:
        raise HTTPException(status_code=400, detail="저장된 학생 결과가 없어 재채점할 수 없습니다")

    new_id = str(uuid.uuid4())
    grading_sessions[new_id] = GradingSession(
        session_id=new_id,
        status="pending",
        progress=0.0,
        total_students=len(students),
        processed_students=0,
        results=[]
    )

    db.add(models.GradingSessionDB(
        id=new_id,
        subject_id=orig.subject_id,
        subject_item_id=orig.subject_item_id,
        user_id=current_user["id"],
        status="running",
        total_students=len(students),
        processed_students=0,
        grading_model=body.grading_model,
        criteria_json=orig.criteria_json,
        answer_problems_json=orig.answer_problems_json,
        regraded_from=orig.regraded_from or orig.id,  # 루트 세션 기준으로 그룹화
    ))
    db.commit()

    background_tasks.add_task(
        run_regrade_session,
        new_id, answer_problems, students, criteria, body.grading_model
    )

    return {"session_id": new_id, "total_students": len(students)}


@app.get("/grading/session/{session_id}/results")
async def get_results(session_id: str, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    session = grading_sessions.get(session_id)
    if session:
        if session.status != "completed":
            raise HTTPException(status_code=400, detail="채점이 아직 완료되지 않았습니다")
        return session.results

    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record or db_record.status != "completed":
        raise HTTPException(status_code=400, detail="채점이 완료되지 않았습니다")

    results = []
    if db_record.results_json:
        results = [StudentResult(**r) for r in json.loads(db_record.results_json)]
    return results


@app.post("/grading/session/{session_id}/cancel")
async def cancel_grading(
    session_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """진행 중인 채점을 강제 중단. 지금까지 채점된 결과는 보존됨."""
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    is_owner = db_record.user_id == current_user["id"]
    is_admin = current_user.get("role") == "admin"
    if not (is_owner or is_admin):
        raise HTTPException(status_code=403, detail="중단 권한이 없습니다")

    if db_record.status not in ("running", "pending"):
        raise HTTPException(status_code=400, detail="진행 중인 채점이 아닙니다")

    # 다음 iteration에서 종료되도록 플래그 설정
    cancelled_sessions.add(session_id)

    # 메모리 세션이 있으면 현재까지 결과를 DB에 즉시 반영
    mem_session = grading_sessions.get(session_id)
    if mem_session:
        results_data = [_dump_strip_images(r.model_dump()) for r in mem_session.results]
        db_record.results_json = json.dumps(results_data, ensure_ascii=False)
        db_record.processed_students = mem_session.processed_students
        db_record.progress = mem_session.progress

    db_record.status = "cancelled"
    db_record.error = "사용자 요청으로 채점이 중단되었습니다"
    db.commit()

    return {"message": "채점이 중단되었습니다", "status": "cancelled"}


@app.delete("/grading/session/{session_id}")
async def delete_session(
    session_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """채점 세션 삭제 (소유자 또는 admin만 가능)."""
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    is_owner = db_record.user_id == current_user["id"]
    is_admin = current_user.get("role") == "admin"
    if not (is_owner or is_admin):
        raise HTTPException(status_code=403, detail="삭제 권한이 없습니다")

    # 진행 중 세션은 삭제 불가 (먼저 강제 중단해야 함)
    if db_record.status == "running":
        raise HTTPException(status_code=400, detail="진행 중인 채점은 먼저 강제 중단 후 삭제할 수 있습니다")

    # FK로 연결된 수정 이력 먼저 삭제
    db.query(models.ProblemRevisionLog).filter(
        models.ProblemRevisionLog.session_id == session_id
    ).delete(synchronize_session=False)

    # 메모리 세션도 정리 (있다면)
    grading_sessions.pop(session_id, None)
    cancelled_sessions.discard(session_id)

    db.delete(db_record)
    db.commit()
    return {"message": "채점 기록이 삭제되었습니다"}


@app.patch("/grading/session/{session_id}/revise")
async def revise_problem_score(
    session_id: str,
    revision: ProblemRevisionRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """교수가 점수 또는 코멘트를 수정. 이력은 자동 기록됨."""
    # 1. 권한 확인 (교수 또는 관리자만)
    if current_user.get("role") not in ("professor", "admin"):
        raise HTTPException(status_code=403, detail="교수 권한이 필요합니다")

    # 2. DB 세션 가져오기
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record or not db_record.results_json:
        raise HTTPException(status_code=404, detail="채점 세션을 찾을 수 없습니다")

    # 3. 세션 소유자 확인 (해당 강의 교수만)
    if db_record.user_id and db_record.user_id != current_user["id"] and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="해당 채점 세션의 수정 권한이 없습니다")

    # 4. 결과 파싱 후 해당 학생/문제 찾기
    results = json.loads(db_record.results_json)
    target_student = None
    target_problem = None
    for student in results:
        if student.get("filename") == revision.student_filename:
            target_student = student
            for p in student.get("problems", []):
                if str(p.get("problem_id")) == str(revision.problem_id):
                    target_problem = p
                    break
            break

    if not target_student or not target_problem:
        raise HTTPException(status_code=404, detail="해당 학생/문제를 찾을 수 없습니다")

    full_score = float(target_problem.get("full_score", 0))
    revision_logs = []

    # 5-A. partial_scores 수정 (각 세부 항목)
    if revision.partial_scores is not None:
        existing_partials = target_problem.get("partial_scores", [])
        for i, new_ps in enumerate(revision.partial_scores):
            if i >= len(existing_partials):
                continue
            old_ps = existing_partials[i]
            max_score = float(old_ps.get("max_score", 0))
            new_score = float(new_ps.score)

            # 점수 범위 검증: 0 ~ max_score
            if new_score < 0 or new_score > max_score:
                raise HTTPException(
                    status_code=400,
                    detail=f"세부 점수는 0 ~ {max_score}점 범위여야 합니다 (입력값: {new_score})"
                )

            # 점수 변경 시 이력 기록
            if old_ps.get("score") != new_score:
                revision_logs.append(models.ProblemRevisionLog(
                    session_id=session_id,
                    student_filename=revision.student_filename,
                    problem_id=str(revision.problem_id),
                    field_name="partial_score",
                    partial_score_index=i,
                    old_value=str(old_ps.get("score")),
                    new_value=str(new_score),
                    revised_by=current_user["id"],
                ))
                existing_partials[i]["score"] = new_score

            # 사유 변경 시 이력 기록
            if new_ps.reason and old_ps.get("reason") != new_ps.reason:
                revision_logs.append(models.ProblemRevisionLog(
                    session_id=session_id,
                    student_filename=revision.student_filename,
                    problem_id=str(revision.problem_id),
                    field_name="partial_reason",
                    partial_score_index=i,
                    old_value=old_ps.get("reason"),
                    new_value=new_ps.reason,
                    revised_by=current_user["id"],
                ))
                existing_partials[i]["reason"] = new_ps.reason

        # 세부 항목 합계로 obtained_score 자동 재계산
        target_problem["obtained_score"] = round(sum(p.get("score", 0) for p in existing_partials), 2)

    # 5-B. obtained_score 직접 수정 (세부 항목이 없는 경우만)
    elif revision.obtained_score is not None:
        if not target_problem.get("partial_scores"):
            new_score = float(revision.obtained_score)
            if new_score < 0 or new_score > full_score:
                raise HTTPException(
                    status_code=400,
                    detail=f"점수는 0 ~ {full_score}점 범위여야 합니다 (입력값: {new_score})"
                )
            old_score = target_problem.get("obtained_score")
            if old_score != new_score:
                revision_logs.append(models.ProblemRevisionLog(
                    session_id=session_id,
                    student_filename=revision.student_filename,
                    problem_id=str(revision.problem_id),
                    field_name="obtained_score",
                    old_value=str(old_score),
                    new_value=str(new_score),
                    revised_by=current_user["id"],
                ))
                target_problem["obtained_score"] = new_score
        else:
            raise HTTPException(
                status_code=400,
                detail="세부 항목이 있는 문제는 obtained_score를 직접 수정할 수 없습니다 (partial_scores를 수정하세요)"
            )

    # 5-C. 교수 코멘트 수정
    if revision.professor_feedback is not None:
        old_feedback = target_problem.get("professor_feedback")
        if old_feedback != revision.professor_feedback:
            revision_logs.append(models.ProblemRevisionLog(
                session_id=session_id,
                student_filename=revision.student_filename,
                problem_id=str(revision.problem_id),
                field_name="professor_feedback",
                old_value=old_feedback,
                new_value=revision.professor_feedback,
                revised_by=current_user["id"],
            ))
            target_problem["professor_feedback"] = revision.professor_feedback

    # 6. 수정 표시 + 학생 총점 재계산
    if revision_logs:
        target_problem["is_revised"] = True
        target_problem["revised_at"] = datetime.utcnow().isoformat()
        # 부분점수 항목 재감지 (0 < score < max_score)
        target_problem["has_partial_score"] = any(
            0 < float(ps.get("score", 0)) < float(ps.get("max_score", 0))
            for ps in target_problem.get("partial_scores", [])
        )
        target_student["total_score"] = round(
            sum(p.get("obtained_score", 0) for p in target_student.get("problems", [])), 2
        )

        # 7. DB 저장
        db_record.results_json = json.dumps(results, ensure_ascii=False)
        for log in revision_logs:
            db.add(log)
        db.commit()

    return {
        "success": True,
        "revisions_count": len(revision_logs),
        "updated_problem": target_problem,
        "updated_total_score": target_student["total_score"]
    }


@app.get("/grading/session/{session_id}/revisions")
async def get_revision_logs(
    session_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """수정 이력 조회"""
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")

    logs = db.query(models.ProblemRevisionLog).filter(
        models.ProblemRevisionLog.session_id == session_id
    ).order_by(models.ProblemRevisionLog.revised_at.desc()).all()

    result = []
    for log in logs:
        user = db.query(models.User).filter(models.User.id == log.revised_by).first()
        result.append({
            "id": log.id,
            "student_filename": log.student_filename,
            "problem_id": log.problem_id,
            "field_name": log.field_name,
            "partial_score_index": log.partial_score_index,
            "old_value": log.old_value,
            "new_value": log.new_value,
            "revised_by_username": user.username if user else None,
            "revised_at": log.revised_at.isoformat(),
        })
    return result


@app.get("/grading/all-revisions")
async def get_all_revision_logs(
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """현재 교수가 본인 채점 세션에서 수행한 모든 수정 이력을 반환 (최신순).
    세션 메타데이터(과목명, 항목명)와 함께 반환."""
    # 본인이 채점한 세션 ID 목록
    my_session_ids = [
        r.id for r in db.query(models.GradingSessionDB)
        .filter(models.GradingSessionDB.user_id == current_user["id"])
        .all()
    ]
    if not my_session_ids:
        return []

    logs = (
        db.query(models.ProblemRevisionLog)
        .filter(models.ProblemRevisionLog.session_id.in_(my_session_ids))
        .order_by(models.ProblemRevisionLog.revised_at.desc())
        .limit(500)
        .all()
    )

    # 세션 메타데이터 캐시
    sessions_map = {
        r.id: r for r in db.query(models.GradingSessionDB)
        .filter(models.GradingSessionDB.id.in_(my_session_ids))
        .all()
    }

    # 사용자 캐시
    user_ids = {log.revised_by for log in logs}
    users_map = {
        u.id: u.username for u in db.query(models.User)
        .filter(models.User.id.in_(user_ids)).all()
    } if user_ids else {}

    # 세부 항목명 캐시
    item_ids = {s.subject_item_id for s in sessions_map.values() if s.subject_item_id}
    items_map = {
        i.id: i.name for i in db.query(models.SubjectItem)
        .filter(models.SubjectItem.id.in_(item_ids)).all()
    } if item_ids else {}

    result = []
    for log in logs:
        sess = sessions_map.get(log.session_id)
        result.append({
            "id": log.id,
            "session_id": log.session_id,
            "subject_name": (sess.subject.name if sess and sess.subject else None),
            "subject_item_name": (items_map.get(sess.subject_item_id) if sess else None),
            "student_filename": log.student_filename,
            "problem_id": log.problem_id,
            "field_name": log.field_name,
            "partial_score_index": log.partial_score_index,
            "old_value": log.old_value,
            "new_value": log.new_value,
            "revised_by_username": users_map.get(log.revised_by),
            "revised_at": log.revised_at.isoformat() if log.revised_at else "",
        })
    return result


@app.get("/grading/session/{session_id}/download")
async def download_excel(
    session_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # DB 레코드 항상 조회 (수정 반영 + 유효성 검사용)
    db_record = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.id == session_id
    ).first()
    if not db_record or db_record.status != "completed":
        raise HTTPException(status_code=400, detail="채점이 완료되지 않았습니다")

    # 원본 AI 점수: 인메모리 우선(수정 전 원점수), 없으면 DB
    mem_session = grading_sessions.get(session_id)
    if mem_session and mem_session.status == "completed":
        original_results = mem_session.results
    elif db_record.results_json:
        original_results = [StudentResult(**r) for r in json.loads(db_record.results_json)]
    else:
        original_results = []

    # 수정 후 점수: 항상 DB (교수 수정사항 반영)
    if db_record.results_json:
        revised_results = [StudentResult(**r) for r in json.loads(db_record.results_json)]
    else:
        revised_results = original_results

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    ai_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    revised_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    highlight_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    # AI 오류 = 노란색 / 부분점수 = 파란색 (UI와 동일한 색상 컨벤션)
    ai_error_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    partial_score_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def get_problem_ids(results):
        ids = []
        for r in results:
            for p in r.problems:
                if p.problem_id not in ids:
                    ids.append(p.problem_id)
        ids.sort()
        return ids

    def build_rank_map(results):
        score_list = [(i, s.total_score) for i, s in enumerate(results)]
        score_list.sort(key=lambda x: x[1], reverse=True)
        rank_map = {}
        for rank, (i, score) in enumerate(score_list, 1):
            if rank > 1 and score == score_list[rank - 2][1]:
                rank_map[i] = rank_map[score_list[rank - 2][0]]
            else:
                rank_map[i] = rank
        return rank_map

    def apply_sheet_style(ws, results, headers):
        for col in range(1, len(headers) + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(results) + 2)),
                default=0
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 4, 12), 50)
        ws.freeze_panes = "C2"

    wb = openpyxl.Workbook()

    # ── Sheet 1: AI 채점결과 (원본) ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "AI채점결과"
    pids = get_problem_ids(original_results)
    headers1 = ["학번", "이름"]
    for pid in pids:
        prob = next((p for r in original_results for p in r.problems if p.problem_id == pid), None)
        pid_str = str(pid) if str(pid).startswith('Q') else f"Q{pid}"
        headers1.append(f"{pid_str} ({prob.full_score if prob else 0}점)")
    headers1.extend(["총점", "순위"])

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = ai_fill
        cell.alignment = center_align
        cell.border = thin_border

    rank_map1 = build_rank_map(original_results)
    sorted_pairs1 = sorted(enumerate(original_results), key=lambda x: rank_map1[x[0]])
    for row_idx, (orig_idx, student) in enumerate(sorted_pairs1, 2):
        ws1.cell(row=row_idx, column=1, value=student.student_id)
        ws1.cell(row=row_idx, column=2, value=student.student_name or "")
        col = 3
        for pid in pids:
            p = next((p for p in student.problems if p.problem_id == pid), None)
            score_cell = ws1.cell(row=row_idx, column=col, value=p.obtained_score if p else 0)
            # AI 오류(노랑) > 부분점수(파랑) 우선순위로 색상 적용
            if p:
                if p.has_ai_error:
                    score_cell.fill = ai_error_fill
                elif getattr(p, "has_partial_score", False):
                    score_cell.fill = partial_score_fill
            col += 1
        ws1.cell(row=row_idx, column=col, value=student.total_score)
        ws1.cell(row=row_idx, column=col + 1, value=rank_map1[orig_idx])
        for c in range(1, len(headers1) + 1):
            ws1.cell(row=row_idx, column=c).border = thin_border
            ws1.cell(row=row_idx, column=c).alignment = center_align

    apply_sheet_style(ws1, original_results, headers1)

    # ── Sheet 2: AI 분석결과 (세부 채점항목) ─────────────────────────────
    ws2 = wb.create_sheet("AI분석결과")
    detail_headers = ["학번", "이름", "문제", "최대점수", "획득점수", "채점항목", "학생답변", "AI피드백", "AI종합피드백"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = ai_fill
        cell.alignment = center_align
        cell.border = thin_border

    row = 2
    for student in original_results:
        for problem in student.problems:
            student_answer = unicodedata.normalize("NFC", "\n\n".join(
                c.source for c in problem.code_cells if c.source.strip()
            )) if problem.code_cells else ""
            for ps_idx, ps in enumerate(problem.partial_scores):
                ws2.cell(row=row, column=1, value=student.student_id).alignment = center_align
                ws2.cell(row=row, column=2, value=student.student_name or "").alignment = center_align
                ws2.cell(row=row, column=3, value=f"Q{problem.problem_id}").alignment = center_align
                ws2.cell(row=row, column=4, value=ps.max_score).alignment = center_align
                score_cell = ws2.cell(row=row, column=5, value=ps.score)
                score_cell.alignment = center_align
                # 획득점수 컬럼 색상: AI 오류(노랑) > 부분점수(파랑) 우선순위
                if problem.has_ai_error:
                    score_cell.fill = ai_error_fill
                elif 0 < float(ps.score) < float(ps.max_score):
                    score_cell.fill = partial_score_fill
                ws2.cell(row=row, column=6, value=ps.item).alignment = wrap_align
                if ps_idx == 0:
                    ws2.cell(row=row, column=7, value=student_answer).alignment = wrap_align
                    ws2.cell(row=row, column=9, value=problem.ai_feedback or "").alignment = wrap_align
                ws2.cell(row=row, column=8, value=ps.reason).alignment = wrap_align
                for c in range(1, 10):
                    ws2.cell(row=row, column=c).border = thin_border
                row += 1

    # 컬럼별 너비: 학번/이름/문제/점수는 좁게, 채점항목/학생답변/피드백은 넓게 고정
    col_widths = {1: 15, 2: 12, 3: 8, 4: 10, 5: 10, 6: 40, 7: 50, 8: 50, 9: 50}
    for col in range(1, 10):
        ws2.column_dimensions[get_column_letter(col)].width = col_widths.get(col, 20)
    ws2.freeze_panes = "C2"

    # ── Sheet 3: 수정후채점결과 (교수 수정점수 + 코멘트) ─────────────────
    ws3 = wb.create_sheet("수정후채점결과")
    pids_rev = get_problem_ids(revised_results)
    headers3 = ["학번", "이름"]
    for pid in pids_rev:
        prob = next((p for r in revised_results for p in r.problems if p.problem_id == pid), None)
        pid_str = str(pid) if str(pid).startswith('Q') else f"Q{pid}"
        headers3.append(f"{pid_str} ({prob.full_score if prob else 0}점)")
        headers3.append(f"{pid_str} 교수코멘트")
    headers3.extend(["총점", "순위"])

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = revised_fill
        cell.alignment = center_align
        cell.border = thin_border

    rank_map3 = build_rank_map(revised_results)
    # 순위 오름차순 정렬 (1위부터 표시)
    sorted_pairs3 = sorted(enumerate(revised_results), key=lambda x: rank_map3[x[0]])
    for row_idx, (orig_idx, student) in enumerate(sorted_pairs3, 2):
        ws3.cell(row=row_idx, column=1, value=student.student_id).alignment = center_align
        ws3.cell(row=row_idx, column=2, value=student.student_name or "").alignment = center_align
        col = 3
        for pid in pids_rev:
            p = next((p for p in student.problems if p.problem_id == pid), None)
            score_cell = ws3.cell(row=row_idx, column=col, value=p.obtained_score if p else 0)
            score_cell.alignment = center_align
            if p and p.is_revised:
                score_cell.fill = highlight_fill
            col += 1
            comment_cell = ws3.cell(row=row_idx, column=col, value=p.professor_feedback if p and p.professor_feedback else "")
            comment_cell.alignment = wrap_align
            col += 1
        ws3.cell(row=row_idx, column=col, value=student.total_score).alignment = center_align
        ws3.cell(row=row_idx, column=col + 1, value=rank_map3[orig_idx]).alignment = center_align
        for c in range(1, len(headers3) + 1):
            ws3.cell(row=row_idx, column=c).border = thin_border

    for col in range(1, len(headers3) + 1):
        max_len = max(
            (len(str(ws3.cell(row=r, column=col).value or "")) for r in range(1, len(revised_results) + 2)),
            default=0
        )
        # 교수코멘트 열은 더 넓게
        is_comment_col = col >= 3 and (col - 3) % 2 == 1 and col < len(headers3) - 1
        ws3.column_dimensions[get_column_letter(col)].width = min(max(max_len + 4, 12), 50 if is_comment_col else 20)
    ws3.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 파일명: 과목이름_세부항목_MMDD.xlsx
    subject = db.query(models.Subject).filter(models.Subject.id == db_record.subject_id).first() if db_record.subject_id else None
    item = db.query(models.SubjectItem).filter(models.SubjectItem.id == db_record.subject_item_id).first() if db_record.subject_item_id else None
    date_src = db_record.completed_at or db_record.created_at
    date_str = date_src.strftime("%m%d") if date_src else session_id[:8]
    subject_name = subject.name if subject else "채점결과"
    item_name = item.name if item else ""
    parts = [p for p in [subject_name, item_name, date_str] if p]
    filename = "_".join(parts) + ".xlsx"
    # 파일명에서 공백·특수문자 제거
    import re
    filename = re.sub(r'[\\/:*?"<>|\s]', '_', filename)

    from urllib.parse import quote
    # RFC 5987 형식: filename*=UTF-8''<URL-encoded-filename>
    encoded_filename = quote(filename, safe='')

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )



@app.post("/rubric/parse-notebook")
async def parse_notebook_rubric(file: UploadFile = File(...)):
    """
    노트북 파일(.ipynb)의 마크다운 셀에서 문제를 읽어 JSON 루브릭으로 변환합니다.

    입력: .ipynb 파일
    - 마크다운 셀에 다음 형식으로 문제 작성:
      ## Q1. 문제 설명 (총점: 4점)
      1. 세부 조건 (1점)
      2. 세부 조건 (점수 표기 없음)

    파싱 결과: evaluation_guideline과 partial_score_criteria로 분류
    """
    try:
        from services.notebook_service import parse_notebook, extract_markdown_cells
        from utils.markdown_parser import parse_markdown_problems

        content = await file.read()
        nb = parse_notebook(content)
        markdown_text = extract_markdown_cells(nb)

        if not markdown_text.strip():
            raise ValueError("노트북에 마크다운 셀이 없거나 비어있습니다.")

        problems, global_guideline = parse_markdown_problems(markdown_text)

        exam_title = file.filename.replace(".ipynb", "").replace("_", " ")

        return {
            "exam_title": exam_title,
            "global_evaluation_guideline": global_guideline,
            "problems": problems
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파싱 오류: {str(e)}")


@app.post("/rubric/parse-markdown")
async def parse_markdown_rubric(file: UploadFile = File(...)):
    """
    마크다운 파일(.md)을 JSON 루브릭으로 변환합니다.

    입력 형식:
    ## Q1. 문제 설명 (총점: 4점)
    1. 세부 조건 (1점)
    2. 세부 조건 (점수 표기 없음)
    """
    try:
        from utils.markdown_parser import parse_markdown_problems

        content = await file.read()
        markdown_text = content.decode('utf-8')

        problems, global_guideline = parse_markdown_problems(markdown_text)

        return {
            "success": True,
            "filename": file.filename,
            "count": len(problems),
            "global_evaluation_guideline": global_guideline,
            "problems": problems
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파싱 오류: {str(e)}")


@app.post("/rubric/parse-markdown-text")
async def parse_markdown_text(text: str):
    """
    마크다운 텍스트를 JSON 루브릭으로 변환합니다. (텍스트 직접 입력)
    """
    try:
        from utils.markdown_parser import parse_markdown_problems

        problems, global_guideline = parse_markdown_problems(text)

        return {
            "success": True,
            "count": len(problems),
            "global_evaluation_guideline": global_guideline,
            "problems": problems
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파싱 오류: {str(e)}")


@app.post("/rubric/decompose-items")
async def decompose_rubric_items(
    request: DecomposeRequest,
    _current_user=Depends(get_current_user)
):
    """루브릭 항목 하나를 동사 단위 세부 항목으로 분해합니다."""
    try:
        decomposed = await decompose_rubric_item_with_ai(
            item=request.item,
            problem_context=request.problem_context or ""
        )
        return {"decomposed_items": decomposed}
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health():
    return {"status": "ok"}


# ─── Admin ────────────────────────────────────────────────────────────────────

@app.get("/admin/stats")
async def admin_stats(
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """시스템 전체 통계."""
    from sqlalchemy import func

    total_users = db.query(models.User).count()
    total_sessions = db.query(models.GradingSessionDB).count()
    completed_sessions = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.status == "completed"
    ).count()
    running_sessions = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.status == "running"
    ).count()

    total_students_graded = db.query(func.sum(models.GradingSessionDB.processed_students)).scalar() or 0
    total_tokens_used = db.query(func.sum(models.GradingSessionDB.tokens_used)).scalar() or 0

    # 최근 7일 세션 수
    from datetime import datetime, timedelta
    week_ago = datetime.utcnow() - timedelta(days=7)
    recent_sessions = db.query(models.GradingSessionDB).filter(
        models.GradingSessionDB.created_at >= week_ago
    ).count()

    # 사용자별 통계
    user_stats = []
    users = db.query(models.User).all()
    for u in users:
        session_count = db.query(models.GradingSessionDB).filter(
            models.GradingSessionDB.user_id == u.id
        ).count()
        subject_count = db.query(models.Subject).filter(
            models.Subject.user_id == u.id
        ).count()
        user_stats.append({
            "id": u.id,
            "username": u.username,
            "role": u.role,
            "sessions": session_count,
            "subjects": subject_count,
        })

    return {
        "total_users": total_users,
        "total_sessions": total_sessions,
        "completed_sessions": completed_sessions,
        "running_sessions": running_sessions,
        "total_students_graded": total_students_graded,
        "total_tokens_used": total_tokens_used,
        "recent_sessions_7d": recent_sessions,
        "user_stats": user_stats,
    }


@app.get("/admin/users")
async def admin_list_users(
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """전체 사용자 목록."""
    users = db.query(models.User).order_by(models.User.created_at).all()
    return [
        {
            "id": u.id,
            "username": u.username,
            "email": u.email,
            "role": u.role,
            "created_at": u.created_at.isoformat() if u.created_at else "",
        }
        for u in users
    ]


@app.put("/admin/users/{user_id}")
async def admin_update_user(
    user_id: int,
    body: dict,
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """사용자 역할 변경 / 비밀번호 초기화."""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    if "role" in body:
        if body["role"] not in ("admin", "professor", "ta", "student"):
            raise HTTPException(status_code=400, detail="유효하지 않은 역할입니다")
        user.role = body["role"]

    if "password" in body:
        if len(body["password"]) < 6:
            raise HTTPException(status_code=400, detail="비밀번호는 6자 이상이어야 합니다")
        user.hashed_password = get_password_hash(body["password"])

    db.commit()
    return {"message": "사용자 정보가 업데이트되었습니다"}


@app.delete("/admin/users/{user_id}")
async def admin_delete_user(
    user_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """사용자 삭제."""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    if user.role == "admin":
        raise HTTPException(status_code=400, detail="관리자 계정은 삭제할 수 없습니다")

    db.delete(user)
    db.commit()
    return {"message": "사용자가 삭제되었습니다"}


@app.get("/admin/settings")
async def admin_get_settings(
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """시스템 설정 조회."""
    settings = db.query(models.SystemSetting).all()
    result = {}
    for s in settings:
        # API 키는 마스킹하여 반환
        if "api_key" in s.key and s.value:
            masked = s.value[:8] + "..." + s.value[-4:] if len(s.value) > 12 else "***"
            result[s.key] = masked
        else:
            result[s.key] = s.value
    return result


@app.put("/admin/settings")
async def admin_update_settings(
    body: dict,
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """시스템 설정 업데이트."""
    allowed_keys = {"openai_api_key", "fireworks_api_key", "llm_model", "base_system_prompt", "max_upload_size_mb"}
    for key, value in body.items():
        if key not in allowed_keys:
            continue
        setting = db.query(models.SystemSetting).filter(models.SystemSetting.key == key).first()
        if setting:
            setting.value = str(value)
        else:
            db.add(models.SystemSetting(key=key, value=str(value)))

        # API 키가 변경되면 환경변수도 즉시 갱신
        if key == "openai_api_key" and value:
            os.environ["OPENAI_API_KEY"] = str(value)
        elif key == "fireworks_api_key" and value:
            os.environ["FIREWORKS_API_KEY"] = str(value)

    db.commit()
    return {"message": "설정이 저장되었습니다"}


@app.get("/admin/sessions")
async def admin_list_sessions(
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """전체 채점 세션 목록 (모든 사용자)."""
    from services.llm_service import AVAILABLE_MODELS
    model_label_map = {m["id"]: m["label"] for m in AVAILABLE_MODELS}

    records = (
        db.query(models.GradingSessionDB)
        .order_by(models.GradingSessionDB.created_at.desc())
        .limit(100)
        .all()
    )
    result = []
    for r in records:
        user = db.query(models.User).filter(models.User.id == r.user_id).first()
        result.append({
            "session_id": r.id,
            "username": user.username if user else "unknown",
            "subject_name": r.subject.name if r.subject else None,
            "status": r.status,
            "total_students": r.total_students,
            "processed_students": r.processed_students,
            "grading_model": r.grading_model,
            "grading_model_label": model_label_map.get(r.grading_model, r.grading_model),
            "created_at": _to_kst(r.created_at),
            "completed_at": _to_kst(r.completed_at) if r.completed_at else None,
        })
    return result


@app.get("/admin/db/schema")
async def admin_db_schema(
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """DB 스키마 정보 반환 (테이블 + 컬럼 목록)."""
    excluded_tables = {"grading_sessions_db"}
    table_descriptions = {
        "users": "교수/관리자 계정",
        "subjects": "수업 정보",
        "subject_items": "수업별 시험/과제",
        "system_settings": "시스템 설정값",
        "problem_revision_logs": "교수 점수/코멘트 수정 이력",
    }
    schema = []
    for table_name, table in models.Base.metadata.tables.items():
        if table_name in excluded_tables:
            continue
        columns = []
        for col in table.columns:
            columns.append({
                "name": col.name,
                "type": str(col.type),
                "nullable": col.nullable,
                "primary_key": col.primary_key,
            })
        schema.append({
            "table": table_name,
            "description": table_descriptions.get(table_name, ""),
            "columns": columns,
        })
    return schema


@app.post("/admin/db/query")
async def admin_db_query(
    body: dict = Body(...),
    admin=Depends(require_admin),
    db: Session = Depends(get_db)
):
    """SELECT 쿼리만 실행 (읽기 전용, 보안)."""
    from sqlalchemy import text
    query = (body.get("query") or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="쿼리가 비어있습니다")

    # 세미콜론 차단 (다중 쿼리 방지: SELECT ...;DROP TABLE ... 같은 공격)
    if ";" in query:
        raise HTTPException(status_code=400, detail="세미콜론(;)은 허용되지 않습니다 (단일 쿼리만 가능)")

    # SELECT 외 차단
    lower = query.lower()
    if not lower.lstrip().startswith("select"):
        raise HTTPException(status_code=400, detail="SELECT 쿼리만 허용됩니다")

    # 금지 테이블 접근 차단
    blocked_tables = ["grading_sessions_db"]
    for tbl in blocked_tables:
        if tbl in lower:
            raise HTTPException(status_code=403, detail=f"'{tbl}' 테이블은 조회할 수 없습니다")

    # 금지 키워드 차단 (공백·세미콜론 기준으로 단어 분리하여 정확히 체크)
    import re
    forbidden = {"drop", "delete", "update", "insert", "alter", "truncate", "create", "grant", "revoke"}
    words = set(re.split(r"[\s,();]+", lower))
    blocked = words & forbidden
    if blocked:
        raise HTTPException(status_code=400, detail=f"'{next(iter(blocked))}' 키워드는 허용되지 않습니다")

    try:
        result = db.execute(text(query))
        rows = result.fetchall()
        columns = list(result.keys())
        # 결과를 직렬화 가능한 형태로 변환
        data = []
        for row in rows[:1000]:  # 최대 1000건
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i]
                if isinstance(val, (datetime,)):
                    val = val.isoformat()
                elif val is not None and not isinstance(val, (str, int, float, bool, list, dict)):
                    val = str(val)
                # 보안: 비밀번호 해시 등 민감 컬럼은 마스킹
                if val is not None and any(word in col.lower() for word in ("password", "token", "secret")):
                    val = "***"
                row_dict[col] = val
            data.append(row_dict)
        return {
            "columns": columns,
            "rows": data,
            "row_count": len(data),
            "truncated": len(rows) > 1000,
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"쿼리 실행 오류: {str(e)}")
